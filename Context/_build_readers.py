# -*- coding: utf-8 -*-
"""Genera doc-previews.js: HTML fiel del contenido de cada documento para el lector in-app."""
import io,sys,json,html
sys.stdout=io.TextIOWrapper(sys.stdout.buffer,encoding='utf-8')
from docx import Document
from docx.oxml.ns import qn
from docx.table import Table
from docx.text.paragraph import Paragraph
import openpyxl

B=r"C:\Users\raulh\Desktop\Claude\PlastHub"
esc=lambda s: html.escape(str(s))

def iter_blocks(doc):
    for child in doc.element.body.iterchildren():
        if child.tag==qn('w:p'): yield ('p',Paragraph(child,doc))
        elif child.tag==qn('w:tbl'): yield ('t',Table(child,doc))

def tbl_html(t):
    rows=t.rows
    out=['<table>']
    for ri,r in enumerate(rows):
        cells=[c.text.strip() for c in r.cells]
        tag='th' if ri==0 else 'td'
        out.append('<tr>'+''.join(f'<{tag}>{esc(c)}</{tag}>' for c in cells)+'</tr>')
    out.append('</table>')
    return ''.join(out)

def docx_html(path):
    doc=Document(path); out=[]; inlist=False
    for kind,obj in iter_blocks(doc):
        if kind=='t':
            if inlist: out.append('</ul>'); inlist=False
            out.append(tbl_html(obj)); continue
        p=obj; txt=p.text.strip()
        if not txt:
            continue
        s=(p.style.name or '') if p.style else ''
        if s.startswith('List'):
            if not inlist: out.append('<ul>'); inlist=True
            out.append(f'<li>{esc(txt)}</li>'); continue
        if inlist: out.append('</ul>'); inlist=False
        if s.startswith('Title'): out.append(f'<h2>{esc(txt)}</h2>')
        elif s.startswith('Heading 1'): out.append(f'<h2>{esc(txt)}</h2>')
        elif s.startswith('Heading 2'): out.append(f'<h3>{esc(txt)}</h3>')
        elif s.startswith('Heading 3'): out.append(f'<h4>{esc(txt)}</h4>')
        else: out.append(f'<p>{esc(txt)}</p>')
    if inlist: out.append('</ul>')
    return ''.join(out)

def xlsx_html(path,sheet,maxcol=12,maxrow=60):
    wb=openpyxl.load_workbook(path,data_only=True); ws=wb[sheet]
    out=['<table>']; rownum=0
    for r in ws.iter_rows(values_only=True):
        cells=list(r)[:maxcol]
        if not any(v is not None and str(v).strip()!='' for v in cells): continue
        rownum+=1
        if rownum>maxrow: break
        tag='th' if rownum==1 else 'td'
        out.append('<tr>'+''.join(f'<{tag}>{esc("" if v is None else v)}</{tag}>' for v in cells)+'</tr>')
    out.append('</table>')
    return ''.join(out)

previews={}
previews['plan']=docx_html(B+r"\PlasHub_Business_Plan.docx")
previews['charter']=docx_html(B+r"\Amaru-I-Project-Charter-v2.docx")
previews['mou']=docx_html(B+r"\MOU_Suministro_Alex_PlasHub_BORRADOR.docx")
# tracker: hoja principal
wbk=openpyxl.load_workbook(B+r"\PlasHub_Permits_Legal_Tracker.xlsx",data_only=True)
previews['tracker']="<h2>Tracker de Permisos y Legal — Ruta crítica</h2>"+xlsx_html(B+r"\PlasHub_Permits_Legal_Tracker.xlsx", wbk.sheetnames[0])

# profile: resumen curado (las celdas son fórmulas; mostramos las cifras clave + corrección)
previews['profile']=(
 "<h2>Modelo Financiero Corregido (Profile v3)</h2>"
 "<p><b>Qué se corrigió:</b> en la pestaña <i>3b. Flujo de Caja</i>, la fila «Pagos a proveedores» estaba "
 "fijada al precio del M1 (−$21,200) para los 12 meses. Ahora usa el precio/flete de cada mes (≈ −$11,360 desde M2).</p>"
 "<table><tr><th>Métrica</th><th>v2 (con error)</th><th>v3 (corregido)</th></tr>"
 "<tr><td>Pagos a proveedores M2–M12</td><td>−$21,200/mes</td><td>−$11,360/mes</td></tr>"
 "<tr><td>Caja acumulada M12</td><td>−$142,958</td><td>−$34,718</td></tr>"
 "<tr><td>Caja mínima (mes)</td><td>—</td><td>−$50,233 (M1)</td></tr></table>"
 "<h3>Supuestos clave (caso base)</h3>"
 "<table><tr><th>Parámetro</th><th>Valor</th></tr>"
 "<tr><td>Volumen</td><td>80 t/mes</td></tr><tr><td>Precio compra ponderado</td><td>$130/t (M2+)</td></tr>"
 "<tr><td>Precio venta ponderado</td><td>$248/t (M2+)</td></tr><tr><td>OPEX fijo</td><td>$8,175/mes</td></tr>"
 "<tr><td>Break-even</td><td>61 t/mes</td></tr><tr><td>EBITDA M2+</td><td>+$305/mes</td></tr>"
 "<tr><td>CAPEX</td><td>$20,858</td></tr></table>"
 "<p style='color:#9a5a05'><b>Nota:</b> abre el archivo en Excel para recalcular las fórmulas con tus propios valores. "
 "Para un modelo totalmente editable usa la herramienta «Plan de Negocio».</p>"
)

js="window.DOCPREVIEWS = "+json.dumps(previews,ensure_ascii=False)+";\n"
open(B+r"\doc-previews.js","w",encoding="utf-8").write(js)
print("OK -> doc-previews.js  ("+str(len(js))+" bytes)")
for k,v in previews.items(): print(f"  {k}: {len(v)} chars")
