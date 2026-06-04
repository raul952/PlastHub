# -*- coding: utf-8 -*-
import io,sys; sys.stdout=io.TextIOWrapper(sys.stdout.buffer,encoding='utf-8')
import openpyxl
p=r"C:\Users\raulh\Desktop\Claude\PlastHub\Context\PlasHub_Project_Profile_v2.xlsx"
wb=openpyxl.load_workbook(p,data_only=False)
print("SHEETS:",wb.sheetnames)
ws=wb["3b. Flujo de Caja"]
print("DIMS",ws.dimensions)
for r in range(1,40):
    label=ws.cell(r,1).value
    if label is None:
        # also check col B for label
        label=ws.cell(r,2).value
    rowvals=[]
    for c in range(1,16):
        v=ws.cell(r,c).value
        if v is not None: rowvals.append(f"{ws.cell(r,c).coordinate}={repr(v)}")
    if rowvals:
        print(f"R{r}: "+" | ".join(rowvals))
